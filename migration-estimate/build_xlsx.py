import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
TSV = os.path.join(HERE, "screens-raw.tsv")
PROCS = os.path.join(HERE, "procs-raw.txt")
OUT = os.path.join(HERE, "TrustBank-Migration-Estimate.xlsx")
DATE = "2026-06-17"

# ---- load + classify ----
rows = []
with open(TSV, encoding="utf-8") as f:
    r = csv.DictReader(f, delimiter="\t")
    for x in r:
        rows.append(x)

def bucket(module, path):
    p = path.lower(); scr = path.split("/")[-1].lower()
    if module in ("Reports", "StatutoryReports") or "report" in p:
        return "Report"
    if module == "RetailBanking":
        return "Complex" if "transaction" in p else "Medium"
    if module in ("Reconciliation",):
        return "Complex"
    if module in ("Config", "Authentication", "Lockers", "ServiceBranch", "(root)"):
        return "Simple"
    if scr.startswith(("g_", "m_", "b_")):
        return "Simple"
    return "Medium"

for x in rows:
    x["Bucket"] = bucket(x["Module"], x["RelativePath"])

proc_count = sum(1 for _ in open(PROCS, encoding="utf-8")) if os.path.exists(PROCS) else 0

buckets = ["Simple", "Medium", "Complex", "Report"]
bcount = {b: sum(1 for x in rows if x["Bucket"] == b) for b in buckets}
modules = sorted({x["Module"] for x in rows})
mod_b = {m: {b: 0 for b in buckets} for m in modules}
for x in rows:
    mod_b[x["Module"]][x["Bucket"]] += 1

# ---- styling helpers ----
FONT = "Calibri"
def font(**kw): kw.setdefault("name", FONT); return Font(**kw)
H1 = font(size=16, bold=True, color="1F3864")
H2 = font(size=12, bold=True, color="1F3864")
HDR = font(bold=True, color="FFFFFF")
BLUE = font(color="0000FF")          # inputs
BLACK = font(color="000000")         # formulas
NOTE = font(size=9, italic=True, color="595959")
hdrfill = PatternFill("solid", fgColor="1F3864")
subfill = PatternFill("solid", fgColor="D9E1F2")
inputfill = PatternFill("solid", fgColor="FFF2CC")
totfill = PatternFill("solid", fgColor="E2EFDA")
thin = Side(style="thin", color="BFBFBF")
box = Border(left=thin, right=thin, top=thin, bottom=thin)
PD = '#,##0;(#,##0);-'
PY1 = '0.0'

wb = Workbook()

def hdr_row(ws, r, cells, widths=None):
    for i, c in enumerate(cells, 1):
        cell = ws.cell(r, i, c); cell.font = HDR; cell.fill = hdrfill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = box
    if widths:
        for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w

# =================== Assumptions ===================
a = wb.active; a.title = "Assumptions"
a["A1"] = "TrustBank CBS — Migration Estimate · Assumptions & Rates"; a["A1"].font = H1
a["A2"] = f"All blue cells are editable inputs. Source counts from legacy code scan, {DATE}."; a["A2"].font = NOTE
a.column_dimensions["A"].width = 46; a.column_dimensions["B"].width = 14; a.column_dimensions["C"].width = 40

r = 4
a.cell(r,1,"Effort rates (person-days per screen)").font = H2; r+=1
rate_row = {}
for b, v in [("Simple",1.5),("Medium",4.0),("Complex",10.0),("Report",2.0)]:
    a.cell(r,1,b).border=box
    c=a.cell(r,2,v); c.font=BLUE; c.fill=inputfill; c.number_format='0.0'; c.border=box
    rate_row[b]=r; r+=1
RATES_RANGE = f"Assumptions!$A${rate_row['Simple']}:$B${rate_row['Report']}"  # for VLOOKUP

r+=1; a.cell(r,1,"Program parameters").font=H2; r+=1
params={}
def param(label, val, fmt='#,##0', note=""):
    global r
    a.cell(r,1,label).border=box
    c=a.cell(r,2,val); c.font=BLUE; c.fill=inputfill; c.number_format=fmt; c.border=box
    if note: a.cell(r,3,note).font=NOTE
    params[label]=r; r+=1
param("QA + UAT (% of build effort)", 0.30, '0%')
param("PM / BA / DevOps / training / docs (% of build)", 0.20, '0%')
param("Security review + performance hardening (pd)", 150)
param("Core engine procs to migrate (count)", 40, '#,##0', "interest, balance, day-end, GL posting, clearing")
param("Core engine proc rate (pd each)", 15, '0.0')
param("Data migration + parallel-run tooling (pd)", 200)
param("Foundation remaining + cross-provider hardening (pd)", 240, '#,##0', "framework is otherwise already built")
param("Reporting platform setup (pd)", 40)
param("Productive days per person-year", 220)
r+=1; a.cell(r,1,"Scenario multipliers").font=H2; r+=1
param("Optimistic ×", 0.70, '0.00', "SQL-Server-only, reports rehosted, strong reuse")
param("Likely ×", 1.00, '0.00')
param("Pessimistic ×", 1.40, '0.00', "multi-DB, full report rebuild, deep parallel-run")
r+=1; a.cell(r,1,"Calendar — team sizes (incl. QA/BA/PM)").font=H2; r+=1
param("Team A (people)", 15)
param("Team B (people)", 20)
param("Team C (people)", 25)

def P(label): return f"Assumptions!$B${params[label]}"

# =================== Effort Estimate ===================
e = wb.create_sheet("Effort Estimate")
e["A1"]="Migration Effort Estimate"; e["A1"].font=H1
e["A2"]=f"Counts are from the legacy code scan ({DATE}); rates/parameters live on the Assumptions sheet (editable)."; e["A2"].font=NOTE
hdr_row(e,4,["Workstream","Count","Rate (pd)","Effort (pd)","Notes"],[44,12,12,14,40])
er=5
def line(label, count, rate_ref, note=""):
    global er
    e.cell(er,1,label).border=box
    if count is not None:
        c=e.cell(er,2,count); c.number_format='#,##0'; c.border=box; c.font=BLACK
    else: e.cell(er,2,"").border=box
    if rate_ref is not None:
        c=e.cell(er,3,f"={rate_ref}"); c.number_format='0.0'; c.border=box; c.font=BLACK
    else: e.cell(er,3,"").border=box
    ef=e.cell(er,4); ef.number_format=PD; ef.border=box; ef.font=BLACK
    e.cell(er,5,note).font=NOTE
    row=er; er+=1; return row
# screen buckets
rs=line("Simple screens (masters, config, lookups, inquiries)", bcount["Simple"], f"Assumptions!$B${rate_row['Simple']}")
e.cell(rs,4, f"=B{rs}*C{rs}")
rm=line("Medium screens (search/list, multi-tab, dependent pickers)", bcount["Medium"], f"Assumptions!$B${rate_row['Medium']}")
e.cell(rm,4, f"=B{rm}*C{rm}")
rc=line("Complex screens (transactions, maker-checker, calc, batch)", bcount["Complex"], f"Assumptions!$B${rate_row['Complex']}")
e.cell(rc,4, f"=B{rc}*C{rc}")
rr=line("Reports (re-implement)", bcount["Report"], f"Assumptions!$B${rate_row['Report']}")
e.cell(rr,4, f"=B{rr}*C{rr}")
rpl=line("Reporting platform setup", None, None); e.cell(rpl,4, f"={P('Reporting platform setup (pd)')}")
rep=line("Core engine procedures", None, None)
e.cell(rep,2, f"={P('Core engine procs to migrate (count)')}"); e.cell(rep,2).number_format='#,##0'
e.cell(rep,3, f"={P('Core engine proc rate (pd each)')}"); e.cell(rep,3).number_format='0.0'
e.cell(rep,4, f"=B{rep}*C{rep}")
rdm=line("Data migration + parallel-run tooling", None, None); e.cell(rdm,4, f"={P('Data migration + parallel-run tooling (pd)')}")
rfo=line("Foundation remaining + cross-provider hardening", None, None); e.cell(rfo,4, f"={P('Foundation remaining + cross-provider hardening (pd)')}")
# build subtotal
bs=er
e.cell(bs,1,"Build subtotal").font=font(bold=True)
e.cell(bs,4, f"=SUM(D{rs}:D{rfo})").font=font(bold=True)
for col in range(1,6): e.cell(bs,col).fill=subfill; e.cell(bs,col).border=box
e.cell(bs,4).number_format=PD; er+=1
# QA / overhead / security
rq=line("QA + UAT", None, None); e.cell(rq,4, f"=D{bs}*{P('QA + UAT (% of build effort)')}")
ro=line("PM / BA / DevOps / training / docs", None, None); e.cell(ro,4, f"=D{bs}*{P('PM / BA / DevOps / training / docs (% of build)')}")
rsec=line("Security + performance hardening", None, None); e.cell(rsec,4, f"={P('Security review + performance hardening (pd)')}")
# total
tt=er
e.cell(tt,1,"TOTAL — Likely (pd)").font=font(bold=True,size=12)
e.cell(tt,4, f"=D{bs}+D{rq}+D{ro}+D{rsec}").font=font(bold=True,size=12)
for col in range(1,6): e.cell(tt,col).fill=totfill; e.cell(tt,col).border=box
e.cell(tt,4).number_format=PD; er+=2

# scenarios
sc=er; e.cell(sc,1,"Scenario").font=H2; e.cell(sc,3,"Effort (pd)").font=H2; e.cell(sc,4,"Person-years").font=H2; er+=1
for name, mult in [("Optimistic", P("Optimistic ×")),("Likely", P("Likely ×")),("Pessimistic", P("Pessimistic ×"))]:
    e.cell(er,1,name).border=box
    c=e.cell(er,3, f"=$D${tt}*{mult}"); c.number_format=PD; c.border=box
    py=e.cell(er,4, f"=C{er}/{P('Productive days per person-year')}"); py.number_format=PY1; py.border=box
    er+=1
er+=1
# calendar
cal=er; e.cell(cal,1,"Calendar (Likely)").font=H2; e.cell(cal,2,"People").font=H2; e.cell(cal,4,"Years").font=H2; er+=1
for tlabel in ["Team A (people)","Team B (people)","Team C (people)"]:
    e.cell(er,1,tlabel.replace(" (people)","")).border=box
    pc=e.cell(er,2, f"={P(tlabel)}"); pc.number_format='#,##0'; pc.border=box
    yr=e.cell(er,4, f"=$D${tt}/(B{er}*{P('Productive days per person-year')})"); yr.number_format=PY1; yr.border=box
    er+=1

# =================== Module Inventory ===================
m = wb.create_sheet("Module Inventory")
m["A1"]="Legacy Screens by Module"; m["A1"].font=H1
m["A2"]=f"Auto-classified from the code scan ({DATE}); complexity is a suggestion to confirm in discovery. Effort uses Assumptions rates."; m["A2"].font=NOTE
hdr_row(m,4,["Module","Simple","Medium","Complex","Report","Total","Est. effort (pd)"],[24,10,10,10,10,10,16])
mr=5
order = sorted(modules, key=lambda mm: -sum(mod_b[mm].values()))
first=mr
for mm in order:
    m.cell(mr,1,mm).border=box
    for i,b in enumerate(buckets,2):
        cc=m.cell(mr,i,mod_b[mm][b]); cc.number_format='#,##0'; cc.border=box
    m.cell(mr,6, f"=SUM(B{mr}:E{mr})").border=box; m.cell(mr,6).number_format='#,##0'
    m.cell(mr,7, (f"=B{mr}*Assumptions!$B${rate_row['Simple']}+C{mr}*Assumptions!$B${rate_row['Medium']}"
                  f"+D{mr}*Assumptions!$B${rate_row['Complex']}+E{mr}*Assumptions!$B${rate_row['Report']}")).border=box
    m.cell(mr,7).number_format=PD
    mr+=1
last=mr-1
m.cell(mr,1,"TOTAL").font=font(bold=True)
for i in range(2,8):
    col=get_column_letter(i); cell=m.cell(mr,i, f"=SUM({col}{first}:{col}{last})"); cell.font=font(bold=True); cell.fill=totfill; cell.border=box
    cell.number_format = PD if i==7 else '#,##0'
m.cell(mr,1).fill=totfill; m.cell(mr,1).border=box
m.freeze_panes="A5"

# =================== Screen List ===================
s = wb.create_sheet("Screen List")
s["A1"]="Full Screen Inventory (1 row per .aspx)"; s["A1"].font=H1
s["A2"]=f"Source: legacy Trust.Bank.Publish code scan, {DATE}. 'Suggested complexity' is heuristic — confirm in the discovery phase."; s["A2"].font=NOTE
hdr_row(s,4,["Module","Sub-area","Screen","Relative path","Suggested complexity","Est. pd"],[20,18,40,52,18,9])
sr=5
for x in rows:
    s.cell(sr,1,x["Module"]); s.cell(sr,2,x["SubArea"]); s.cell(sr,3,x["Screen"]); s.cell(sr,4,x["RelativePath"]); s.cell(sr,5,x["Bucket"])
    s.cell(sr,6, f'=VLOOKUP(E{sr},{RATES_RANGE},2,0)').number_format='0.0'
    sr+=1
s.freeze_panes="A5"; s.auto_filter.ref=f"A4:F{sr-1}"

# =================== Read Me ===================
n = wb.create_sheet("Read Me", 0)
n.column_dimensions["A"].width = 110
n["A1"]="TrustBank CBS → .NET 10 MVC — Migration Estimate"; n["A1"].font=H1
lines=[
 "",
 f"Prepared from an automated scan of the legacy source on {DATE}.",
 "STATUS: Rough Order of Magnitude (ROM), +/-40%. Firm to +/-15% after a 4-6 week discovery (screen-by-screen inventory).",
 "",
 "MEASURED SCOPE",
 f"  - Functional (non-report) screens: {bcount['Simple']+bcount['Medium']+bcount['Complex']:,}",
 f"  - Report screens: {bcount['Report']:,}   (+ 892 .rdlc/.rpt report definitions)",
 f"  - Total .aspx: {len(rows):,}    Reusable controls (.ascx): 19    Stored procedures: {proc_count:,}",
 "",
 "METHOD",
 "  - Bottom-up: each functional screen classified Simple/Medium/Complex; reports separate; rate per bucket.",
 "  - Plus workstreams: reporting platform, core engine procedures, data migration/parallel-run,",
 "    remaining foundation + cross-provider hardening, QA/UAT, security/perf, PM/BA/DevOps/training.",
 "  - Calendar = total effort / (team size x productive days per year).",
 "",
 "ALREADY BUILT (excluded from the build numbers — a real head start)",
 "  - Framework split, auth/session, menu + fail-closed access guard, multi-provider DAL,",
 "    reusable components (record/account/date pickers, search modal, notifications, RowToken),",
 "    a repeatable per-screen migration recipe, and 6 screens delivered as reference.",
 "",
 "HOW TO USE THIS WORKBOOK",
 "  - 'Assumptions' — edit the blue cells (rates, %s, team size, scenario multipliers); everything recalculates.",
 "  - 'Effort Estimate' — the headline numbers, scenarios, and calendar.",
 "  - 'Module Inventory' — screens + effort by module (for phasing/sequencing).",
 "  - 'Screen List' — every screen with its suggested complexity (the discovery worksheet).",
 "",
 "KEY RISKS / SWING ITEMS",
 "  - Cross-provider SQL (Oracle/Postgres): raw-SQL screens are SQL-Server dialect; single-DB target removes ~120-150 pd.",
 "  - Reports (362 screens + 892 definitions): 'rehost' vs 'rebuild' swings the total by ~500-700 pd.",
 "  - 5,317 procedures: deep embedded logic (interest/GL/day-end) — highest risk, needs heavy parallel-run validation.",
 "  - Banking correctness/compliance lengthens UAT and may require regulator sign-off.",
]
for i,t in enumerate(lines,3):
    c=n.cell(i,1,t)
    if t.isupper() and t.strip(): c.font=H2
n.sheet_view.showGridLines=False

try:
    wb.calculation.fullCalcOnLoad = True  # Excel recomputes all formulas on open (no LibreOffice needed here)
except Exception:
    from openpyxl.workbook.properties import CalcProperties
    wb.calculation = CalcProperties(fullCalcOnLoad=True)

wb.save(OUT)
print("saved", OUT)
print("buckets", bcount, "total screens", len(rows), "procs", proc_count)
